"""
Generate Excel Bank Statement Import File for Odoo
====================================================
Creates a professionally formatted Excel file that can be
imported into Odoo via Accounting > Bank > Import (Upload).

Odoo expects these columns for bank statement import:
  - Date
  - Label (payment reference / memo)
  - Partner (optional - customer/vendor name)
  - Amount
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import datetime
import os

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "Bank_Statement_Import.xlsx")

# ──────────────────────────────────────────────────────────
# BANK TRANSACTIONS DATA (Realistic client bank statement)
# ──────────────────────────────────────────────────────────
# Using dates spread across a month for realism
base = datetime.date(2026, 2, 1)

transactions = [
    # ── Week 1: Opening + first payments ──
    (base, "Opening Balance Transfer", "", 10000.00),
    (base + datetime.timedelta(1), "NEFT-CR-TechSol-INV001", "Tech Solutions Inc.", 1500.00),
    (base + datetime.timedelta(1), "NEFT-CR-GlobalMfg-SOFT22", "Global Manufacturing Co.", 500.00),
    (base + datetime.timedelta(2), "CHQ DEP 100234 - RetailStores", "Retail Stores Group", 1200.00),
    (base + datetime.timedelta(3), "UPI/CR/4521789/ABCEnt", "ABC Enterprises", 900.00),
    (base + datetime.timedelta(3), "RTGS-CR-QuickBuy-MAINT", "Quick Buy Ltd", 2350.00),

    # ── Week 2: Vendor payments + fees ──
    (base + datetime.timedelta(7), "NEFT-DR-OfficeSup-PO1001", "Office Supplies Co.", -450.00),
    (base + datetime.timedelta(7), "RTGS-DR-TechHW-PO1002", "Tech Hardware Supplier", -2400.00),
    (base + datetime.timedelta(8), "AUTO-DR-CloudSvc-SUB", "Cloud Services LLC", -49.99),
    (base + datetime.timedelta(8), "NEFT-DR-EquipRent-MNTH", "Equipment Rentals Inc.", -303.00),
    (base + datetime.timedelta(9), "CHQ 001001 - RawMat Supplier", "Raw Materials Supplier", -1800.00),
    (base + datetime.timedelta(9), "BANK SERVICE CHARGE - FEB", "", -15.00),
    (base + datetime.timedelta(9), "INTEREST CREDIT - FEB", "", 12.50),

    # ── Week 3: Partial payments, multi-invoice ──
    (base + datetime.timedelta(14), "NEFT-CR-TechSol-PARTIAL", "Tech Solutions Inc.", 950.00),
    (base + datetime.timedelta(14), "UPI/CR/8834521/TechSol-MULTI", "Tech Solutions Inc.", 600.00),
    (base + datetime.timedelta(15), "WIRE TRANSFER FEE - INTL", "", -35.00),
    (base + datetime.timedelta(15), "CC PROCESSING FEE", "", -8.75),

    # ── Week 4: Special scenarios ──
    (base + datetime.timedelta(20), "NEFT-CR-RetailStr-ADVANCE", "Retail Stores Group", 1500.00),
    (base + datetime.timedelta(21), "DEPOSIT - UNKNOWN SOURCE", "", 250.00),
    (base + datetime.timedelta(22), "CR-REFUND-RawMat-RET", "Raw Materials Supplier", 200.00),
    (base + datetime.timedelta(22), "ATM WDL - PETTY CASH", "", -500.00),

    # ── BONUS: Extra transactions for more practice ──
    (base + datetime.timedelta(23), "NEFT-CR-ABCEnt-DEPOSIT", "ABC Enterprises", 750.00),
    (base + datetime.timedelta(24), "SALARY PAYMENT - JAN 2026", "", -3500.00),
    (base + datetime.timedelta(24), "ELECTRICITY BILL - FEB", "", -275.00),
    (base + datetime.timedelta(25), "INSURANCE PREMIUM - Q1", "", -1200.00),
    (base + datetime.timedelta(25), "UPI/CR/9912345/CASH-SALE", "", 480.00),
    (base + datetime.timedelta(26), "TAX REFUND - IT DEPT", "", 890.00),
    (base + datetime.timedelta(27), "NEFT-DR-CloudSvc-ANNUAL", "Cloud Services LLC", -599.88),
    (base + datetime.timedelta(27), "RENT PAYMENT - FEB 2026", "", -2000.00),
    (base + datetime.timedelta(28), "MISC CREDIT - ROUNDING ADJ", "", 0.12),
]

# ──────────────────────────────────────────────────────────
# CREATE EXCEL WORKBOOK
# ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════
# SHEET 1: IMPORT-READY (Odoo format)
# ════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Bank Statement"

# Styles
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='714B67', end_color='714B67', fill_type='solid')  # Odoo purple
data_font = Font(name='Calibri', size=11)
money_format = '#,##0.00'
date_format = 'YYYY-MM-DD'
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
green_font = Font(name='Calibri', size=11, color='006100')
red_font = Font(name='Calibri', size=11, color='9C0006')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Column headers (Odoo standard import columns)
headers = ["Date", "Label", "Partner", "Amount"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data rows
for row_idx, (date, label, partner, amount) in enumerate(transactions, 2):
    # Date
    cell_date = ws.cell(row=row_idx, column=1, value=date)
    cell_date.number_format = date_format
    cell_date.font = data_font
    cell_date.border = thin_border
    cell_date.alignment = Alignment(horizontal='center')

    # Label
    cell_label = ws.cell(row=row_idx, column=2, value=label)
    cell_label.font = data_font
    cell_label.border = thin_border

    # Partner
    cell_partner = ws.cell(row=row_idx, column=3, value=partner)
    cell_partner.font = data_font
    cell_partner.border = thin_border

    # Amount (with conditional formatting)
    cell_amount = ws.cell(row=row_idx, column=4, value=amount)
    cell_amount.number_format = money_format
    cell_amount.border = thin_border
    cell_amount.alignment = Alignment(horizontal='right')
    if amount >= 0:
        cell_amount.font = green_font
        cell_amount.fill = green_fill
    else:
        cell_amount.font = red_font
        cell_amount.fill = red_fill

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 16

# Freeze header row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f"A1:D{len(transactions) + 1}"

# ════════════════════════════════════════════════════════════
# SHEET 2: RECONCILIATION CHEAT SHEET
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Reconciliation Guide")

guide_header_font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
guide_fill = PatternFill(start_color='714B67', end_color='714B67', fill_type='solid')
section_font = Font(name='Calibri', bold=True, size=11, color='714B67')
note_font = Font(name='Calibri', size=10, color='666666', italic=True)

guide_data = [
    ("Row", "Transaction", "Amount", "Scenario Type", "How to Reconcile"),
    (2, "Opening Balance Transfer", "$10,000.00", "OPENING BALANCE",
     "Use 'Manual Operations' tab → Set counterpart account (e.g., Owner's Equity / Capital)"),
    (3, "NEFT-CR-TechSol-INV001", "$1,500.00", "AUTO-MATCH ✅",
     "Odoo should suggest INV/25-26/0001 automatically → Click Validate"),
    (4, "NEFT-CR-GlobalMfg-SOFT22", "$500.00", "AUTO-MATCH ✅",
     "Odoo should suggest INV/25-26/0002 automatically → Click Validate"),
    (5, "CHQ DEP 100234 - RetailStores", "$1,200.00", "AUTO-MATCH ✅",
     "Matches INV/25-26/0003 → Click Validate"),
    (6, "UPI/CR-ABCEnt", "$900.00", "MANUAL MATCH 🔍",
     "Search for INV/25-26/0004 in Match Existing Entries tab → Select → Validate"),
    (7, "RTGS-CR-QuickBuy-MAINT", "$2,350.00", "WRITE-OFF (Bank Fee) 📝",
     "Match INV/25-26/0005 ($2,400) → $50 difference → Add write-off line for Bank Fees → Validate"),
    (8, "NEFT-DR-OfficeSup-PO1001", "-$450.00", "AUTO-MATCH ✅",
     "Matches BILL/25-26/02/0001 → Click Validate"),
    (9, "RTGS-DR-TechHW-PO1002", "-$2,400.00", "AUTO-MATCH ✅",
     "Matches BILL/25-26/02/0002 → Click Validate"),
    (10, "AUTO-DR-CloudSvc-SUB", "-$49.99", "RECURRING MODEL 🔄",
     "Match BILL/25-26/02/0003 → Then create a Reconciliation Model for future months"),
    (11, "NEFT-DR-EquipRent-MNTH", "-$303.00", "WRITE-OFF (Fee) 📝",
     "Match BILL/25-26/02/0004 ($300) → $3 diff = convenience fee → Write-off to Bank Charges"),
    (12, "CHQ 001001 - RawMat", "-$1,800.00", "AUTO-MATCH ✅",
     "Matches BILL/25-26/02/0005 → Click Validate"),
    (13, "BANK SERVICE CHARGE", "-$15.00", "WRITE-OFF 📝",
     "No invoice to match → Use Manual Operations → Set account to Bank Fees/Charges"),
    (14, "INTEREST CREDIT", "$12.50", "WRITE-OFF 📝",
     "No invoice → Use Manual Operations → Set account to Interest Income"),
    (15, "NEFT-CR-TechSol-PARTIAL", "$950.00", "PARTIAL PAYMENT ⚡",
     "Match INV/25-26/0006 ($1,000) → Mark as partial → $50 remains open on invoice"),
    (16, "UPI/CR-TechSol-MULTI", "$600.00", "MULTI-INVOICE 🔗",
     "Select BOTH INV/25-26/0007 ($300) + INV/25-26/0008 ($300) → Validate together"),
    (17, "WIRE TRANSFER FEE", "-$35.00", "WRITE-OFF 📝",
     "Manual Operations → Bank Charges account"),
    (18, "CC PROCESSING FEE", "-$8.75", "WRITE-OFF 📝",
     "Manual Operations → Processing Fees / Bank Charges account"),
    (19, "NEFT-CR-RetailStr-ADVANCE", "$1,500.00", "OVERPAYMENT / ADVANCE 💰",
     "No matching invoice → Create payment on customer's receivable account → Will offset future invoices"),
    (20, "DEPOSIT - UNKNOWN", "$250.00", "UNKNOWN DEPOSIT ❓",
     "Set partner first → Then check if invoice exists → If not, use Manual Operations with suspense account"),
    (21, "CR-REFUND-RawMat", "$200.00", "VENDOR REFUND ↩️",
     "This is incoming from a vendor → Match against credit note or use Manual Operations → AP account"),
    (22, "ATM WDL - PETTY CASH", "-$500.00", "CASH TRANSFER 💵",
     "Manual Operations → Set account to Petty Cash / Cash on Hand"),
    (23, "NEFT-CR-ABCEnt-DEPOSIT", "$750.00", "ADVANCE DEPOSIT 💰",
     "No matching invoice → Record as customer advance on Receivables"),
    (24, "SALARY PAYMENT", "-$3,500.00", "EXPENSE 💼",
     "Manual Operations → Set account to Salary Expense"),
    (25, "ELECTRICITY BILL", "-$275.00", "EXPENSE 💼",
     "Manual Operations → Set account to Utilities Expense"),
    (26, "INSURANCE PREMIUM", "-$1,200.00", "PREPAID EXPENSE 📋",
     "Manual Operations → Set account to Prepaid Insurance (asset) or Insurance Expense"),
    (27, "UPI/CR - CASH SALE", "$480.00", "CASH SALE 🛒",
     "Manual Operations → Set account to Sales Revenue (if no invoice created)"),
    (28, "TAX REFUND", "$890.00", "TAX CREDIT 🏛️",
     "Manual Operations → Set account to Tax Receivable / Tax Refund account"),
    (29, "NEFT-DR-CloudSvc-ANNUAL", "-$599.88", "NO BILL YET ⏳",
     "Create vendor bill first, then match → Or use Manual Operations to AP account"),
    (30, "RENT PAYMENT", "-$2,000.00", "EXPENSE 💼",
     "Manual Operations → Set account to Rent Expense"),
    (31, "MISC CREDIT - ROUNDING", "$0.12", "ROUNDING ADJ 🔢",
     "Manual Operations → Set account to Rounding Adjustment / Misc Income"),
]

# Headers
for col, header in enumerate(guide_data[0], 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = guide_header_font
    cell.fill = guide_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data
scenario_colors = {
    "AUTO-MATCH": PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    "MANUAL MATCH": PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    "WRITE-OFF": PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    "PARTIAL": PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    "MULTI-INVOICE": PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
    "OVERPAYMENT": PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    "UNKNOWN": PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),
    "VENDOR REFUND": PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
    "OPENING": PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    "EXPENSE": PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    "RECURRING": PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    "CASH": PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),
    "ADVANCE": PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    "PREPAID": PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    "TAX": PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
    "NO BILL": PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),
    "ROUNDING": PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid'),
    "CASH SALE": PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
}

for row_idx, row_data in enumerate(guide_data[1:], 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # Color-code by scenario type
        scenario = str(row_data[3])
        for key, fill in scenario_colors.items():
            if key in scenario:
                cell.fill = fill
                break

# Column widths for guide sheet
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 32
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 24
ws2.column_dimensions['E'].width = 70

ws2.freeze_panes = 'A2'

# ════════════════════════════════════════════════════════════
# SHEET 3: IMPORT INSTRUCTIONS
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Import Instructions")

instructions = [
    ("HOW TO IMPORT THIS BANK STATEMENT INTO ODOO", ""),
    ("", ""),
    ("Step", "Instructions"),
    ("1", "Go to Accounting module in Odoo"),
    ("2", "Click on your Bank journal card on the Dashboard"),
    ("3", "Click the ⬆️ Upload button (or Import under Transactions)"),
    ("4", "Select this Excel file (Bank_Statement_Import.xlsx)"),
    ("5", "Odoo will show a preview - map columns if needed:"),
    ("", "   • Date → Date"),
    ("", "   • Label → Label"),
    ("", "   • Partner → Partner (optional)"),
    ("", "   • Amount → Amount"),
    ("6", "Click 'Import' to load all transactions"),
    ("7", "You'll see all 30 transactions in your bank journal"),
    ("8", "Click 'Reconcile' to start matching!"),
    ("", ""),
    ("TIPS", ""),
    ("•", "The 'Bank Statement' sheet is the one Odoo imports"),
    ("•", "The 'Reconciliation Guide' sheet tells you how to handle each transaction"),
    ("•", "Green amounts = incoming (customer payments, deposits)"),
    ("•", "Red amounts = outgoing (vendor payments, expenses, fees)"),
    ("•", "Partner column helps Odoo auto-match with the right invoices/bills"),
    ("", ""),
    ("COLUMN MAPPING", ""),
    ("Date", "Transaction date (YYYY-MM-DD format)"),
    ("Label", "Bank reference / memo - used for matching"),
    ("Partner", "Customer or vendor name - must match exactly in Odoo"),
    ("Amount", "Positive = money in, Negative = money out"),
    ("", ""),
    ("NOTE", "Make sure the partners (customers/vendors) already exist in Odoo"),
    ("", "before importing. They were created by the setup script."),
]

for row_idx, (col1, col2) in enumerate(instructions, 1):
    cell1 = ws3.cell(row=row_idx, column=1, value=col1)
    cell2 = ws3.cell(row=row_idx, column=2, value=col2)
    
    if row_idx == 1:
        cell1.font = Font(name='Calibri', bold=True, size=14, color='714B67')
        ws3.merge_cells('A1:B1')
    elif col1 in ("Step", "TIPS", "COLUMN MAPPING", "NOTE"):
        cell1.font = Font(name='Calibri', bold=True, size=11, color='714B67')
        cell2.font = Font(name='Calibri', bold=True, size=11, color='714B67')
    elif col1 and col1[0].isdigit():
        cell1.font = Font(name='Calibri', bold=True, size=11)
        cell2.font = Font(name='Calibri', size=11)
    else:
        cell1.font = Font(name='Calibri', size=11)
        cell2.font = Font(name='Calibri', size=11)

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 70

# ──────────────────────────────────────────────────────────
# SAVE
# ──────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"✅ Excel file created: {OUTPUT_FILE}")
print(f"   📊 Sheet 1: 'Bank Statement' - {len(transactions)} transactions (IMPORT THIS)")
print(f"   📖 Sheet 2: 'Reconciliation Guide' - How to reconcile each row")
print(f"   📝 Sheet 3: 'Import Instructions' - Step-by-step import guide")
print(f"\n🎯 Import the file in Odoo: Accounting > Bank > Upload")
